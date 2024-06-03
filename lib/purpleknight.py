import argparse
import glob
import openpyxl
import os
import lib.logs as logging

@logging.log_call
def get_risk_ids_from_purpleknight_file(file_path:str):

        matching_xlsx_files = glob.glob(file_path)
        if matching_xlsx_files:
            file_path = matching_xlsx_files[0]
            if len(matching_xlsx_files) == 1:
                print(f'PurpleKnight XLSX report found at "{file_path}"')
                print()
            elif len(matching_xlsx_files) > 1:
                print(f'PurpleKnight XLSX report at "{file_path}" selected out of multiple options:')
                print("\n".join([f'-> {matching_xlsx_file}' for matching_xlsx_file in matching_xlsx_files]))
                print()
        else:
            print(f'Impossible to find a PurpleKnight XLSX report at "{file_path}".')
            raise FileNotFoundError

        # IDS
        if ".xlsx" in file_path:
            try:
                risk_ids = get_risk_ids_from_purpleknight_xlsx_file(file_path)
            except FileNotFoundError as e:
                return None
        else:
            print("The PurpleKnight file needs a \".xlsx\" extention")
            return None

        return risk_ids

@logging.log_call
def get_risk_ids_from_purpleknight_xlsx_file(file_path:str) -> list:

    matching_xlsx_files = glob.glob(file_path)
    if matching_xlsx_files:
        file_path = matching_xlsx_files[0]
        if len(matching_xlsx_files) == 1:
            print(f'PurpleKnight XLSX report found at "{file_path}"')
            print()
        elif len(matching_xlsx_files) > 1:
            print(f'PurpleKnight XLSX report at "{file_path}" selected out of multiple options:')
            print("\n".join([f'-> {matching_xlsx_file}' for matching_xlsx_file in matching_xlsx_files]))
            print()
    else:
        print(f'Impossible to find a PurpleKnight XLSX report at "{file_path}".')
        raise FileNotFoundError

    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Select the specified sheet
    sheet = workbook["Indicators results"]
    
    # Extract data from the specified column
    risk_id_values = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=5):
        data_cell = row[0]  # Column B
        condition_cell = row[3]  # Column E
        if condition_cell.value == "IOE Found":
            risk_id_values.append(data_cell.value)
    
    print(f'Risk ids from the PurpleKnight XLSX report at "{file_path}":\n{", ".join(risk_id_values)}')
    print()
    return risk_id_values

@logging.log_call
def request_purpleknight_file_path() -> str:
    file_path = None
    while not file_path:
        try:
            file_path = input("Path to the PurpleKnight XLSX file (Ctrl+C to quit) : ")
        except KeyboardInterrupt as e:
            raise KeyboardInterrupt
        if not os.path.isfile(file_path):
            print(f'Error : File not found at "{file_path}".')
            file_path = None
    return file_path

@logging.log_call
def main():

    # ARGUMENTS
    parser = argparse.ArgumentParser(description='Parse a PurpleKnight XLSX report and extract the list of the risks ID')
    parser.add_argument('-f', '--file', type=str, help='Path to a PurpleKnight XLSX.')
    args = parser.parse_args()

    if hasattr(args, 'file') and args.file is not None:
        file_path = args.file 
    else:
        try:
            file_path = request_purpleknight_file_path()
        except KeyboardInterrupt:
            return
    # IDS
    return get_risk_ids_from_purpleknight_file(file_path)

if __name__ == '__main__':
    main()